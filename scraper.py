#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TikTok Comments Scraper Engine for web-cmt
"""

import sys
import os
import re
import time
import datetime
import urllib.parse
import subprocess
import atexit
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Đảm bảo mã UTF-8 trên Windows
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

# Thư viện HTTP
try:
    from curl_cffi import requests as http_client
    USE_CURL_CFFI = True
except ImportError:
    import requests as http_client
    USE_CURL_CFFI = False

DEFAULT_SERVER_URL = os.environ.get("SERVER_URL", "http://localhost:8080")
DEVICE_ID = "7520531026079925774"
_SERVER_PROCESS = None


def extract_video_id(input_str: str) -> str:
    input_str = input_str.strip()
    if input_str.isdigit():
        return input_str

    match = re.search(r"/video/(\d+)", input_str)
    if match:
        return match.group(1)

    match_digits = re.search(r"(\d{15,22})", input_str)
    if match_digits:
        return match_digits.group(1)

    raise ValueError(f"Không tìm thấy Video ID hợp lệ từ: {input_str}")


def check_server_health(server_url: str = DEFAULT_SERVER_URL) -> bool:
    try:
        resp = http_client.get(f"{server_url}/health", timeout=3)
        if resp.status_code == 200:
            data = resp.json()
            return data.get("status") == "ok" and data.get("ready") is True
    except Exception:
        pass
    return False


def ensure_server_running(server_url: str = DEFAULT_SERVER_URL):
    global _SERVER_PROCESS
    if check_server_health(server_url):
        return

    base_dir = os.path.dirname(os.path.abspath(__file__))
    candidate_paths = [
        os.path.join(base_dir, "server.mjs"),
        os.path.join(base_dir, "..", "tiktok-signature", "server.mjs"),
        os.path.join(os.path.dirname(base_dir), "server.mjs"),
    ]

    server_script = None
    server_cwd = None
    for p in candidate_paths:
        p = os.path.abspath(p)
        if os.path.exists(p):
            server_script = p
            server_cwd = os.path.dirname(p)
            break

    if not server_script:
        return

    creationflags = 0
    if sys.platform == "win32":
        creationflags = subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0x08000000

    try:
        r = http_client.get(f"{server_url}/health", timeout=1)
        if r.status_code == 200:
            pass  # server is currently initializing, just wait
    except Exception:
        node_env = os.environ.copy()
        node_env["PORT"] = "8080"
        node_env["SIGNATURE_PORT"] = "8080"
        _SERVER_PROCESS = subprocess.Popen(
            ["node", server_script],
            cwd=server_cwd,
            env=node_env,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=creationflags
        )

    for _ in range(160):
        time.sleep(0.5)
        if check_server_health(server_url):
            return


def cleanup_server():
    global _SERVER_PROCESS
    if _SERVER_PROCESS:
        try:
            _SERVER_PROCESS.terminate()
            _SERVER_PROCESS.kill()
        except Exception:
            pass


atexit.register(cleanup_server)


def build_comments_url(video_id: str, cursor: int = 0, count: int = 20) -> str:
    params = {
        "aid": "1988",
        "app_language": "en",
        "app_name": "tiktok_web",
        "aweme_id": str(video_id),
        "count": str(count),
        "cursor": str(cursor)
    }
    return f"https://www.tiktok.com/api/comment/list/?{urllib.parse.urlencode(params)}"


def build_replies_url(video_id: str, comment_id: str, cursor: int = 0, count: int = 20) -> str:
    params = {
        "aid": "1988",
        "app_language": "en",
        "app_name": "tiktok_web",
        "item_id": str(video_id),
        "comment_id": str(comment_id),
        "count": str(count),
        "cursor": str(cursor)
    }
    return f"https://www.tiktok.com/api/comment/list/reply/?{urllib.parse.urlencode(params)}"


def fetch_tiktok_api(target_url: str, server_url: str = DEFAULT_SERVER_URL) -> dict:
    try:
        sign_resp = http_client.post(
            f"{server_url}/signature",
            json={"url": target_url},
            headers={"Content-Type": "application/json"},
            timeout=15,
        )
        if sign_resp.status_code == 200:
            sign_data = sign_resp.json()
            if sign_data.get("status") == "ok":
                data_info = sign_data.get("data") or {}
                signed_url = data_info.get("signed_url")
                nav = data_info.get("navigator") or {}
                user_agent = nav.get("user_agent", "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.6 Safari/605.1.15")
                cookies = data_info.get("cookies", "")

                req_headers = {
                    "User-Agent": user_agent,
                    "Accept": "application/json, text/plain, */*",
                    "Referer": "https://www.tiktok.com/",
                    "Sec-Fetch-Dest": "empty",
                    "Sec-Fetch-Mode": "cors",
                    "Sec-Fetch-Site": "same-origin",
                }
                if cookies:
                    req_headers["Cookie"] = cookies

                req_kwargs = {
                    "headers": req_headers,
                    "timeout": 15,
                }
                if USE_CURL_CFFI:
                    req_kwargs["impersonate"] = "safari15_5"

                api_resp = http_client.get(signed_url, **req_kwargs)
                if api_resp.status_code == 200 and api_resp.text.strip():
                    json_res = api_resp.json()
                    if isinstance(json_res, dict) and "comments" in json_res:
                        return json_res
    except Exception:
        pass

    try:
        fetch_resp = http_client.post(
            f"{server_url}/fetch",
            json={"url": target_url},
            headers={"Content-Type": "application/json"},
            timeout=30,
        )
        if fetch_resp.status_code == 200:
            res_json = fetch_resp.json()
            if res_json.get("status") == "ok":
                return res_json.get("data") or {}
    except Exception:
        pass

    return {}


def parse_comment_item(item: dict, parent_comment_id: str = None) -> dict:
    user = item.get("user") or {}
    create_time_ts = item.get("create_time", 0)
    if create_time_ts:
        create_time_str = datetime.datetime.fromtimestamp(
            create_time_ts, datetime.timezone.utc
        ).astimezone().strftime("%Y-%m-%d %H:%M:%S")
    else:
        create_time_str = ""

    avatar_url = ""
    avatar_thumb = user.get("avatar_thumb")
    if isinstance(avatar_thumb, dict):
        url_list = avatar_thumb.get("url_list") or []
        if url_list:
            avatar_url = url_list[0]
    elif isinstance(avatar_thumb, str):
        avatar_url = avatar_thumb

    return {
        "Comment ID": str(item.get("cid", "")),
        "Loại": "Trả lời (Reply)" if parent_comment_id else "Bình luận gốc",
        "Parent Comment ID": parent_comment_id or "",
        "Username": user.get("unique_id") or user.get("uniqueId") or "",
        "Tên hiển thị": user.get("nickname", ""),
        "Nội dung bình luận": item.get("text", ""),
        "Lượt thích (Likes)": item.get("digg_count", 0),
        "Số câu trả lời": item.get("reply_comment_total", 0),
        "Thời gian đăng": create_time_str,
        "Ngôn ngữ": item.get("comment_language", ""),
        "Tác giả thả tim": "Có" if (item.get("is_author_digged") or item.get("is_author_liked") or item.get("author_pin")) else "Không",
        "Ghim đầu trang": "Có" if item.get("is_pinned") else "Không",
        "User ID": str(user.get("uid", "")),
        "Link Avatar": avatar_url,
    }


def scrape_comments_generator(video_id: str, max_comments: int = 0, fetch_replies: bool = True, server_url: str = DEFAULT_SERVER_URL):
    """
    Generator yield tiến độ và bình luận theo thời gian thực cho Web App.
    """
    ensure_server_running(server_url)

    all_comments = []
    seen_cids = set()
    cursor = 0
    page_num = 1
    has_more = True

    yield {
        "type": "log",
        "message": f"🚀 Bắt đầu quét bình luận cho Video ID: {video_id}...",
        "count": 0,
        "progress": 5
    }

    while has_more:
        yield {
            "type": "log",
            "message": f"📄 Đang tải trang {page_num} (cursor = {cursor})...",
            "count": len(all_comments),
            "progress": min(10 + page_num * 15, 85)
        }

        url = build_comments_url(video_id, cursor=cursor, count=50)
        data = fetch_tiktok_api(url, server_url=server_url)
        if not isinstance(data, dict):
            data = {}

        comments_list = data.get("comments") or []
        if not comments_list and page_num == 1:
            # Thử lại thêm 1 lần nữa sau 2s đề phòng server vừa khởi động
            time.sleep(2.0)
            data = fetch_tiktok_api(url, server_url=server_url)
            if isinstance(data, dict):
                comments_list = data.get("comments") or []

        if not comments_list:
            yield {
                "type": "log",
                "message": "ℹ️ Đã lấy hết toàn bộ bình luận từ TikTok.",
                "count": len(all_comments),
                "progress": 90
            }
            break

        new_items = []
        for item in comments_list:
            cid = str(item.get("cid", ""))
            if cid and cid not in seen_cids:
                seen_cids.add(cid)
                parsed = parse_comment_item(item)
                all_comments.append(parsed)
                new_items.append(parsed)

                if max_comments > 0 and len(all_comments) >= max_comments:
                    break

                # Lấy replies
                reply_total = item.get("reply_comment_total", 0)
                if fetch_replies and reply_total > 0:
                    reply_cursor = 0
                    reply_has_more = True
                    while reply_has_more:
                        reply_url = build_replies_url(video_id, comment_id=cid, cursor=reply_cursor, count=50)
                        reply_data = fetch_tiktok_api(reply_url, server_url=server_url)
                        if not isinstance(reply_data, dict):
                            reply_data = {}

                        sub_list = reply_data.get("comments") or []
                        if not sub_list:
                            break

                        for r_item in sub_list:
                            r_cid = str(r_item.get("cid", ""))
                            if r_cid and r_cid not in seen_cids:
                                seen_cids.add(r_cid)
                                r_parsed = parse_comment_item(r_item, parent_comment_id=cid)
                                all_comments.append(r_parsed)
                                new_items.append(r_parsed)
                                if max_comments > 0 and len(all_comments) >= max_comments:
                                    break

                        if max_comments > 0 and len(all_comments) >= max_comments:
                            break

                        reply_has_more = bool(reply_data.get("has_more", 0))
                        next_r_cursor = reply_data.get("cursor", 0)
                        if not reply_has_more or next_r_cursor == reply_cursor:
                            break
                        reply_cursor = next_r_cursor
                        time.sleep(0.2)

        yield {
            "type": "data",
            "new_comments": new_items,
            "total_count": len(all_comments),
            "message": f"✅ Đã tải thêm {len(new_items)} cmt | Tổng: {len(all_comments)} cmt",
            "progress": min(10 + page_num * 20, 90)
        }

        if max_comments > 0 and len(all_comments) >= max_comments:
            break

        has_more = bool(data.get("has_more", 0))
        next_cursor = data.get("cursor", 0)

        if not has_more or next_cursor == cursor:
            break

        cursor = next_cursor
        page_num += 1
        time.sleep(0.5)

    yield {
        "type": "finished",
        "comments": all_comments,
        "total_count": len(all_comments),
        "progress": 100,
        "message": f"🎉 Hoàn thành! Thu thập thành công {len(all_comments)} bình luận."
    }


def save_to_excel(comments_data: list, output_filepath: str):
    if not comments_data:
        return

    df = pd.DataFrame(comments_data)

    # Loại bỏ các cột không cần xuất ra file Excel (SecUid, Link Avatar)
    cols_to_drop = [c for c in ["SecUid", "Link Avatar"] if c in df.columns]
    if cols_to_drop:
        df = df.drop(columns=cols_to_drop)

    df.insert(0, "STT", range(1, len(df) + 1))

    with pd.ExcelWriter(output_filepath, engine="openpyxl") as writer:
        sheet_name = "Bình luận TikTok"
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        worksheet.freeze_panes = "A2"

        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        border_side = Side(style="thin", color="E5E7EB")
        cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = Border(
                left=Side(style="thin", color="374151"),
                right=Side(style="thin", color="374151"),
                top=Side(style="thin", color="374151"),
                bottom=Side(style="medium", color="111827"),
            )
        worksheet.row_dimensions[1].height = 30

        center_cols = {
            "STT", "Comment ID", "Parent Comment ID", "Loại", "Lượt thích (Likes)",
            "Số câu trả lời", "Thời gian đăng", "Ngôn ngữ", "Tác giả thả tim",
            "Ghim đầu trang", "User ID"
        }

        for row_idx in range(2, len(df) + 2):
            worksheet.row_dimensions[row_idx].height = 24
            is_even = (row_idx % 2 == 0)
            row_fill = PatternFill(start_color="F9FAFB" if is_even else "FFFFFF", fill_type="solid")

            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = cell_border
                cell.fill = row_fill
                cell.font = Font(name="Calibri", size=10)

                col_name = df.columns[col_idx - 1]
                if col_name in center_cols:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_name == "Nội dung bình luận":
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        for col_idx, col in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            if col == "Nội dung bình luận":
                worksheet.column_dimensions[col_letter].width = 50
            elif col == "Thời gian đăng":
                worksheet.column_dimensions[col_letter].width = 20
            elif col in ["Username", "Tên hiển thị"]:
                worksheet.column_dimensions[col_letter].width = 22
            else:
                max_len = max(
                    len(str(col)),
                    max((len(str(val or "")) for val in df[col].astype(str)), default=0)
                )
                worksheet.column_dimensions[col_letter].width = max(min(max_len + 4, 35), 10)
